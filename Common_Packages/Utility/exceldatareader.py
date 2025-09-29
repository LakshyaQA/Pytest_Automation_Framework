import openpyxl
from Common_Packages.resources.ConfigPath import UploadFile



class Exceldatareader:

    @staticmethod
    def get_Data_from_excel(sheetname):
        final_list = []
        workbook = openpyxl.load_workbook(UploadFile.file_upload_path('Data_File.xlsx'))
        sheet = workbook[sheetname]
        total_rows = sheet.max_row
        total_cols = sheet.max_column

        for r in range(2, total_rows + 1):
            row_list = []
            for c in range(1, total_cols + 1):
                row_list.append(sheet.cell(row=r, column=c).value)
            final_list.append(row_list)
        print(final_list)
        return final_list

    import openpyxl
    @staticmethod


    def read_column_from_excel( header_name):
        """
        Reads the first value from a specific column in an Excel file based on the header name using openpyxl.

        Parameters:
            file_path (str): Path to the Excel file.
            header_name (str): Name of the header in the Excel file.

        Returns:
            str: Value from the specified column.
        """
        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(UploadFile.file_upload_path('Data_File.xlsx'))
            sheet = wb.active

            # Find the header row
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value == header_name:
                        header_column = cell.column_letter
                        break
                else:
                    continue
                break
            else:
                raise ValueError(f"Header '{header_name}' not found in the Excel file.")

            # Read the first value from the specified column
            first_value = None
            for cell in sheet[header_column][1:]:
                if cell.value is not None:
                    first_value = cell.value
                    break

            if first_value is None:
                raise ValueError(f"No value found in column '{header_name}'.")

            return first_value
        except Exception as e:
            print(f"An error occurred: {e}")
            return None

